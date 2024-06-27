# FAÇO ISSO PARA LER/CARREGAR A "PAGE" DE UMA PLANILHA EXCEL
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("data/pivot_table.xlsx")
sheet = wb["Relatorio"] #ESPECIFICO O NOME DA "PAGE"/WORKBOOK DA MINHA PLANILHA

# AKI EU PEGO AS REFÊRENCIAS ENTRE AS PLANILHAS (MAX E MIN) DAS LINHAS E COLUNAS
min_column = wb.active.min_column
max_column = wb.active.max_column

min_row = wb.active.min_row
max_row = wb.active.max_row

sheet["B6"] = "=SUM(B2:B5)"
sheet["B6"].style = "Currency"

for i in range(min_column+1, max_column+1):
    letter = get_column_letter(i)
    sheet[f"{letter}{max_row+1}"] = f"=SUM({letter}{min_row+1}:{letter}{max_row})"
    sheet[f"{letter}{max_row+1}"].style = "Currency"
    # print(letter)

wb.save("teste.xlsx")