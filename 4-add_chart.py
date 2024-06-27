# FAÇO ISSO PARA LER/CARREGAR A "PAGE" DE UMA PLANILHA EXCEL
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

wb = load_workbook("data/pivot_table.xlsx")
sheet = wb["Relatorio"] #ESPECIFICO O NOME DA "PAGE"/WORKBOOK DA MINHA PLANILHA

# AKI EU PEGO AS REFÊRENCIAS ENTRE AS PLANILHAS (MAX E MIN) DAS LINHAS E COLUNAS
min_column = wb.active.min_column
max_column = wb.active.max_column

# consolezao
# print(min_column)
# print(max_column)

min_row = wb.active.min_row
max_row = wb.active.max_row

# consolezao
# print(min_row)
# print(max_row)

# ---------------------------------------------------------------------------------------
# FIM DA REFERENCIA (LINHA MAX E COLUNA MAX para dev o grafico)
# ---------------------------------------------------------------------------------------

# ADD O GRAFICO
barchart = BarChart()
data = Reference(
    sheet,
    min_col=min_column + 1,
    max_col=max_column,
    min_row=min_row,
    max_row=max_row
)

categories = Reference(
    sheet,
    min_col=min_column,
    max_col=min_column,
    min_row=min_row + 1,
    max_row=max_row
)

# 
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

# CRIANDO O GRAFICO
sheet.add_chart(barchart, "B10")
barchart.title = "Vendas por Fabricantes"
barchart.style = 2

wb.save("data/barchart.xlsx")
